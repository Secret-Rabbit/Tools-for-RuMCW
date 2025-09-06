# Конвертирует таблицу xlsx в словарь json.
# Первый столбец - ключ, второй столбец - значение

from openpyxl import load_workbook
import json

# Пути к файлам
xlsxFilePath = 'Redirects.xlsx'
jsonFilePath = 'Redirects.json'

# Загружаем Excel-файл
wb = load_workbook(xlsxFilePath)
ws = wb.active

# Словарь, который будем формировать
data_dict = {}

# Проходим по строкам, начиная со второй (первая — заголовки или не используется)
for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
    key, value = row
    if key is not None:
        data_dict[str(key)] = value

# Сохраняем в JSON
with open(jsonFilePath, 'w', encoding='utf-8') as f:
    json.dump(data_dict, f, indent=4, ensure_ascii=False)